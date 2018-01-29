"""Common models for capturing exchange, index, security, and historical data

model security with
	name (symbol)
	headers
	table
operations:
	fromAV (initialize)
	fromAV (incremental)
	fromXLSX
	toXLSX

"""

import os
import datetime
import numpy
import openpyxl
from matplotlib import pyplot
from scrapefe import alphavantage as av

class HistDat(object):
    """
    """

    def __init__(self, symbol):
        """
        """
        self.symbol = symbol
        self.header = tuple()
        self.table = numpy.array([])

    def initialize(self):
        """Generate an initial dataset from full pull
        """
        self.header, self.table = av.getHistory(self.symbol)

    def increment(self):
        """Update existing dataset with recent entries
        """
        pass

    def toXLSX(self, xlsxPath):
        """Save existing datset to xlsx worksheet. Sheet name is security
           symbol.
        """
        isNew = False
        if os.path.isfile(xlsxPath):
            wb = openpyxl.load_workbook(xlsxPath)
        else:
            wb = openpyxl.Workbook()
            isNew = True
        if self.symbol in wb.get_sheet_names():
            ws = wb.get_sheet_by_name(self.symbol)
        else:
            ws = wb.create_sheet(self.symbol)
            if isNew:
                wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
        for jj, field in enumerate(self.header):
            ws.cell(row=1, column=jj+1).value = field
        for ii in range(self.table.shape[0]):
            for jj in range(self.table.shape[1]):
                value = self.table[ii,jj]
                if jj is self.header.index('timestamp'):
                    value = datetime.datetime.fromtimestamp(value)
                ws.cell(row=ii+2, column=jj+1).value = value
        wb.save(xlsxPath)

    def plot(self):
        """
        """
        hf = pyplot.figure()
        hx = hf.add_subplot(111)
        ti_nx = self.table[:,0]
        plotFields_ndcs = [1,2,3,4,5]   
        lgnStrs = []
        for ndx in plotFields_ndcs:
            hx.plot(ti_nx, self.table[:,ndx])
            lgnStrs.append(self.header[ndx])
        xTicks = hx.get_xticks()
        xTickLabels = [datetime.datetime.fromtimestamp(tick).strftime('%U/%m/%d') for tick in xTicks]
        hx.legend(lgnStrs)
        hx.set_xticklabels(xTickLabels)
        hx.set_title('Historical Data for %s' % self.symbol)
        hx.set_xlabel('Date')
        hx.set_ylabel('Price')
        return hf

    @classmethod
    def fromXLSX(cls, xlsxPath, sheetName):
        """Load existing dataset from xlsx worksheet
        """
        obj = cls(sheetName)
        wb = openpyxl.load_workbook(xlsxPath)
        ws = wb.get_sheet_by_name(sheetName)
        obj.header = []
        for jj in range(ws.max_column):
            obj.header.append(ws.cell(row=1, column=jj+1).value)
        obj.table = numpy.array([])
        for ii in range(1,ws.max_row):
            row = []
            for jj in range(ws.max_column):
                row.append(ws.cell(row=ii+1, column=jj+1).value)
            row = numpy.array(row)
            if ii is 1:
                obj.table = numpy.append(obj.table, row).reshape(1, -1)
            else:
                obj.table = numpy.append(obj.table, row.reshape(1, -1), axis=0)
        return obj

def test():
    """
    """
    #xlsxPath = __file__.replace('.py', '.xlsx')
    #sq = HistDat('SQ')
    #sq.initialize()
    #sq.toXLSX(xlsxPath)
    sq = HistDat.fromXLSX('test.xlsx', 'SQ')

if __name__ == "__main__":
    test()
